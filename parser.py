from bs4 import BeautifulSoup as bs
from bs4 import Tag
import lxml
import requests
import openpyxl


URL = 'https://www.allscrabblewords.com/{number}-letter-words/'


def parser(url: str):
    req = requests.get(url)
    if req.status_code != 200:
        raise ConnectionError
    html = req.text
    tree = lxml.etree.HTML(html)
    data = tree.xpath('/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()')
    return data


def main():
    i = 2
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    while i <= 12:
        wb.create_sheet(title='page {number}'.format(number=i), index=0)
        sheet = wb['page {number}'.format(number=i)]
        url = URL.format(number=i)
        word = parser(url)
        for j in word:
            cell = sheet.cell(row=word.index(j) + 1, column=1)
            cell.value = j
        i += 1
        print('page {number} parsed'.format(number=i))
    wb.save('words.xlsx')



if __name__ == '__main__':
    main()